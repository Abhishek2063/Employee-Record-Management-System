import csv
import pandas as pd
from fpdf import FPDF
from datetime import datetime
import os
from pathlib import Path
from fastapi import HTTPException

EXPORT_DIR = Path("exports")
EXPORT_DIR.mkdir(exist_ok=True)

def get_export_filename(prefix: str, extension: str):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return EXPORT_DIR / f"{prefix}_{timestamp}.{extension}"

def export_to_csv(data, filename):
    if not data:
        raise HTTPException(status_code=404, detail="No data to export.")
    
    with open(filename, mode="w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        # Headers without User ID and is_first_entry flag
        headers = ["Date", "Name", "Email", "Punch In", "Punch Out", "Duration (hours)", "Total Hours"]
        writer.writerow(headers)
        
        for row in data:
            # Skip the helper flag when writing to CSV
            csv_row = [
                row.get("Date", ""),
                row.get("Name", ""),
                row.get("Email", ""),
                row.get("Punch In", ""),
                row.get("Punch Out", ""),
                row.get("Duration (hours)", ""),
                row.get("Total Hours", "")
            ]
            writer.writerow(csv_row)
    return filename


def export_to_excel(data, filename):
    if not data:
        raise HTTPException(status_code=404, detail="No data to export.")

    # Remove helper flag before creating DataFrame
    clean_data = []
    for row in data:
        clean_row = {k: v for k, v in row.items() if k != "is_first_entry"}
        clean_data.append(clean_row)
    
    df = pd.DataFrame(clean_data)
    
    # Create Excel file with formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Attendance Report', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Attendance Report']
        
        # Add formatting
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add borders and alignment for data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                if row > 1:  # Data rows
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return filename


def export_to_pdf(data, filename):
    if not data:
        raise HTTPException(status_code=404, detail="No data to export.")

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Title
    pdf.set_font("Arial", "B", 16)
    pdf.set_text_color(54, 96, 146)  # Professional blue color
    pdf.cell(0, 10, "Attendance Report", ln=True, align="C")
    pdf.ln(5)
    
    # Date info
    pdf.set_font("Arial", "", 10)
    pdf.set_text_color(0, 0, 0)
    if data and data[0].get("Date"):
        report_date = data[0]["Date"] if data[0]["Date"] else "Multiple Dates"
        pdf.cell(0, 6, f"Report Date: {report_date}", ln=True)
        pdf.cell(0, 6, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
    pdf.ln(5)

    # Table headers
    headers = ["Date", "Name", "Email", "Punch In", "Punch Out", "Duration (hrs)", "Total Hours"]
    page_width = pdf.w - 2 * pdf.l_margin
    col_widths = [25, 35, 50, 25, 25, 25, 25]  # Adjusted widths
    line_height = 8

    # Header styling
    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(54, 96, 146)  # Professional blue
    pdf.set_text_color(255, 255, 255)  # White text

    # Draw header row
    y_start = pdf.get_y()
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], line_height, header, border=1, align="C", fill=True)
    pdf.ln()

    # Data rows
    pdf.set_font("Arial", "", 8)
    pdf.set_text_color(0, 0, 0)
    
    current_user = None
    for i, row in enumerate(data):
        # Clean row data
        clean_row = {k: v for k, v in row.items() if k != "is_first_entry"}
        
        # Alternate row colors for better readability
        if row.get("is_first_entry", True):
            if i % 2 == 0:
                pdf.set_fill_color(245, 245, 245)  # Light gray
            else:
                pdf.set_fill_color(255, 255, 255)  # White
            current_user = row.get("Name", "")
        else:
            # Use same color as previous row for merged appearance
            pass
        
        # Draw cells
        row_data = [
            clean_row.get("Date", ""),
            clean_row.get("Name", ""),
            clean_row.get("Email", ""),
            clean_row.get("Punch In", ""),
            clean_row.get("Punch Out", ""),
            str(clean_row.get("Duration (hours)", "")),
            str(clean_row.get("Total Hours", ""))
        ]
        
        for j, cell_data in enumerate(row_data):
            pdf.cell(col_widths[j], line_height, str(cell_data), border=1, align="C", fill=True)
        pdf.ln()

    pdf.output(str(filename))
    return filename